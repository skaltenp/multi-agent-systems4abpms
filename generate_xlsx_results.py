"""
Generate XLSX evaluation report from experiment results.

Sheet 1 ("Detailed"): Three tables (Completion %, Avg Time, Avg Steps)
  - Rows: Model x Modality
  - Columns: 7 process adaptations (base experiment) + 2 exception columns + Avg
  - Grouped: Operational | Tactical | Exceptions

Sheet 2 ("Summary"): One table
  - Rows: Model x Modality
  - Columns: Overall Completion %, Time avg +/- std, Steps avg +/- std
"""

import json
import csv
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
RESULTS_DIR = BASE_DIR / "experiment_results"
SESSIONS_DIR = BASE_DIR / "sessions"
OUTPUT_DIR = BASE_DIR / "results"

# --- Column definitions ---

# Operational adaptations (base experiment type)
OPERATIONAL = [
    ("base_rule", "Base Rule"),
    ("0_values", "Area 0"),
    ("500_values", "Area 500"),
    ("900_values", "Area 900"),
    ("city_values", "Excl. Cities"),
]

# Tactical adaptations (base experiment type)
TACTICAL = [
    ("extension_estimates", "Mand. Reading"),
    ("extension_mail", "Dir. Mail"),
]

# Exception experiment types (aggregated across all adaptations)
EXCEPTIONS = [
    ("exception_handling", "Typing Err."),
    ("exception_handling_db_error", "DB Err."),
]

ALL_BASE_ADAPTATIONS = [key for key, _ in OPERATIONAL + TACTICAL]

# --- Method config ---

METHOD_CONFIG = {
    "add": {
        "display": "Modular Agents Add Rule",
        "json_files": [
            "frame_agent_process_addition.json",
            "frame_agent_process_execution.json",
            "process_agent.json",
        ],
    },
    "generate_bpmn": {
        "display": "Modular Agents BPMN Rule",
        "json_files": [
            "frame_agent_process_addition.json",
            "frame_agent_process_execution.json",
            "process_agent.json",
        ],
    },
    "classic": {
        "display": "Classic Agent Prompt",
        "json_files": [
            "classic_agent.json",
        ],
    },
}

MODEL_DISPLAY = {
    "gpt-5.4": "GPT-5.4",
    "gpt-5.1": "GPT-5.1",
}


# --- Data loading helpers ---


def discover_models() -> list[str]:
    """Find all model directories that have result CSVs."""
    models = set()
    for f in RESULTS_DIR.iterdir():
        if not (f.name.startswith("process_adaptation_results_") and f.suffix == ".csv"):
            continue
        name = f.stem.replace("process_adaptation_results_", "")
        for method in ["generate_bpmn", "classic", "add"]:
            idx = name.find(f"_{method}")
            if idx > 0:
                models.add(name[:idx])
                break
    return sorted(models)


def load_result_csvs(model_dir: str) -> pd.DataFrame:
    """Load and concatenate all result CSVs for a given model, tagging experiment_type."""
    frames = []
    for f in sorted(RESULTS_DIR.iterdir()):
        if not (f.name.startswith(f"process_adaptation_results_{model_dir}_") and f.suffix == ".csv"):
            continue
        df = pd.read_csv(f)
        name = f.stem.replace(f"process_adaptation_results_{model_dir}_", "")

        # Parse method and experiment type from filename
        if name.startswith("generate_bpmn"):
            method = "generate_bpmn"
            rest = name.replace("generate_bpmn_", "", 1) if "_" in name[len("generate_bpmn"):] else ""
        else:
            parts = name.split("_", 1)
            method = parts[0]
            rest = parts[1] if len(parts) > 1 else ""

        if rest in ("", method):
            experiment_type = "base"
        elif rest == "exception_handling":
            experiment_type = "exception_handling"
        elif rest == "exception_handling_db_error":
            experiment_type = "exception_handling_db_error"
        else:
            experiment_type = "base"

        df["experiment_type"] = experiment_type
        frames.append(df)

    if not frames:
        raise FileNotFoundError(f"No result CSVs found for model '{model_dir}'")
    return pd.concat(frames, ignore_index=True)


def read_timings(session_path: Path) -> dict:
    timings_file = session_path / "timings.csv"
    result = {}
    if timings_file.exists():
        with open(timings_file) as f:
            for row in csv.DictReader(f):
                result[row["agent"]] = float(row["time"])
    return result


def get_total_time(timings: dict, method: str) -> float | None:
    if method in ("add", "generate_bpmn"):
        f, p = timings.get("frame"), timings.get("process")
        return (f + p) if f is not None and p is not None else None
    return timings.get("classic")


def count_ai_messages(session_path: Path, method: str) -> int | None:
    total = 0
    for jf in METHOD_CONFIG[method]["json_files"]:
        jpath = session_path / jf
        if not jpath.exists():
            return None
        try:
            with open(jpath) as f:
                data = json.load(f)
            msgs = data.get("messages", data) if isinstance(data, dict) else data
            total += sum(
                1 for m in msgs
                if isinstance(m, dict) and m.get("id", [None])[-1] == "AIMessage"
            )
        except (json.JSONDecodeError, KeyError):
            return None
    return total


def enrich_with_session_data(df: pd.DataFrame, model_dir: str) -> pd.DataFrame:
    """Add total_time_s and ai_steps columns from session files."""
    times, ai_steps = [], []
    for _, row in df.iterrows():
        sp = SESSIONS_DIR / model_dir / row["session_id"]
        method = row["rule_adaptation_method"]
        timings = read_timings(sp)
        times.append(get_total_time(timings, method))
        ai_steps.append(count_ai_messages(sp, method))
    df = df.copy()
    df["total_time_s"] = times
    df["ai_steps"] = ai_steps
    return df


# --- Report building ---


def _completion_pct(subset: pd.DataFrame) -> float | None:
    if len(subset) == 0:
        return None
    return round(subset["all_correct"].sum() / len(subset) * 100, 1)


def _mean_or_none(series: pd.Series) -> float | None:
    valid = series.dropna()
    return round(valid.mean(), 1) if len(valid) > 0 else None


def _std_or_none(series: pd.Series) -> float | None:
    valid = series.dropna()
    return round(valid.std(), 1) if len(valid) > 1 else None


def _fmt_mean_std(mean_val, std_val) -> str | None:
    if mean_val is None:
        return None
    if std_val is None or np.isnan(std_val):
        return f"{mean_val}"
    return f"{mean_val} \u00b1 {std_val}"


def build_detailed_table(df: pd.DataFrame, model_dir: str, metric: str) -> list[dict]:
    """Build rows for one metric table on the Detailed sheet.
    metric: 'completion', 'time', or 'steps'
    """
    display_model = MODEL_DISPLAY.get(model_dir, model_dir)
    rows = []

    for method, cfg in METHOD_CONFIG.items():
        method_df = df[df["rule_adaptation_method"] == method]
        if method_df.empty:
            continue

        row = {"Model / Modality": f"{display_model} {cfg['display']}"}

        # Operational + Tactical columns: base experiment type, per adaptation
        values_for_avg = []
        for adapt_key, adapt_label in OPERATIONAL + TACTICAL:
            subset = method_df[
                (method_df["process_adaptation"] == adapt_key)
                & (method_df["experiment_type"] == "base")
            ]
            if metric == "completion":
                val = _completion_pct(subset)
            elif metric == "time":
                val = _mean_or_none(subset["total_time_s"])
            else:  # steps
                val = _mean_or_none(subset["ai_steps"])
            row[adapt_label] = val
            if val is not None:
                values_for_avg.append(val)

        # Exception columns: aggregated across all adaptations for that experiment type
        for exp_key, exp_label in EXCEPTIONS:
            subset = method_df[method_df["experiment_type"] == exp_key]
            if metric == "completion":
                val = _completion_pct(subset)
            elif metric == "time":
                val = _mean_or_none(subset["total_time_s"])
            else:
                val = _mean_or_none(subset["ai_steps"])
            row[exp_label] = val
            if val is not None:
                values_for_avg.append(val)

        # Avg column
        row["Avg."] = round(np.mean(values_for_avg), 1) if values_for_avg else None
        rows.append(row)

    return rows


def build_summary_table(df: pd.DataFrame, model_dir: str) -> list[dict]:
    """Build rows for the Summary sheet."""
    display_model = MODEL_DISPLAY.get(model_dir, model_dir)
    rows = []

    for method, cfg in METHOD_CONFIG.items():
        method_df = df[df["rule_adaptation_method"] == method]
        if method_df.empty:
            continue

        row = {"Model / Modality": f"{display_model} {cfg['display']}"}

        # Overall completion across ALL experiment types and adaptations
        total = len(method_df)
        correct = method_df["all_correct"].sum()
        row["Completion (%)"] = round(correct / total * 100, 1) if total > 0 else None

        # Time
        t_mean = _mean_or_none(method_df["total_time_s"])
        t_std = _std_or_none(method_df["total_time_s"])
        row["Avg Time (s)"] = _fmt_mean_std(t_mean, t_std)

        # Steps
        s_mean = _mean_or_none(method_df["ai_steps"])
        s_std = _std_or_none(method_df["ai_steps"])
        row["Avg Steps"] = _fmt_mean_std(s_mean, s_std)

        rows.append(row)

    return rows


# --- XLSX formatting ---

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
GROUP_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
GROUP_FONT = Font(bold=True, size=10)
TABLE_TITLE_FONT = Font(bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
GROUP_RIGHT_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="medium"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def auto_width(ws, df: pd.DataFrame, col_offset: int = 0):
    """Set column widths based on content."""
    for col_idx, col in enumerate(df.columns, 1):
        values = [len(str(v)) for v in df.iloc[:, col_idx - 1] if pd.notna(v)]
        max_len = max(len(str(col)), *values) if values else len(str(col))
        letter = get_column_letter(col_idx + col_offset)
        ws.column_dimensions[letter].width = max_len + 3


def _apply_bold_underline(ws, data_rows: list[dict], columns: list[str],
                          start_row: int, col_offset: int):
    """Bold the best value and underline second best per column (for completion: higher is better)."""
    for col_name in columns:
        col_idx = list(data_rows[0].keys()).index(col_name) + 1 + col_offset
        vals = []
        for i, row in enumerate(data_rows):
            v = row.get(col_name)
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                vals.append((v, i))

        if len(vals) < 2:
            continue

        vals.sort(key=lambda x: x[0], reverse=True)
        best_val, best_idx = vals[0]
        second_val, second_idx = vals[1]

        # Bold best
        cell = ws.cell(row=start_row + best_idx, column=col_idx)
        cell.font = Font(bold=True, size=10)

        # Underline second best
        cell = ws.cell(row=start_row + second_idx, column=col_idx)
        cell.font = Font(underline="single", size=10)


def write_detailed_sheet(ws, all_models_data: dict):
    """Write three tables (Completion, Time, Steps) vertically on the Detailed sheet."""

    metric_configs = [
        ("completion", "Table 3: Completion Rate (%)"),
        ("time", "Table A1: Average Time (s)"),
        ("steps", "Table A2: Average Steps (AI Messages)"),
    ]

    # Column structure for group headers
    # Col A = Model/Modality, then Operational (5), Tactical (2), Exceptions (2), Avg (1)
    group_spans = [
        ("Operational", 2, 6),      # columns B-F (5 cols)
        ("Tactical", 7, 8),         # columns G-H (2 cols)
        ("Exceptions", 9, 10),      # columns I-J (2 cols)
    ]

    current_row = 1

    for metric, title in metric_configs:
        # Table title
        ws.cell(row=current_row, column=1, value=title).font = TABLE_TITLE_FONT
        current_row += 1

        # Group header row
        ws.cell(row=current_row, column=1, value="")
        for group_name, start_col, end_col in group_spans:
            ws.merge_cells(
                start_row=current_row, start_column=start_col,
                end_row=current_row, end_column=end_col
            )
            cell = ws.cell(row=current_row, column=start_col, value=group_name)
            cell.font = GROUP_FONT
            cell.fill = GROUP_FILL
            cell.alignment = Alignment(horizontal="center")
            # Fill all merged cells
            for c in range(start_col, end_col + 1):
                ws.cell(row=current_row, column=c).fill = GROUP_FILL
                ws.cell(row=current_row, column=c).border = THIN_BORDER
        current_row += 1

        # Collect rows from all models
        all_rows = []
        for model_dir in all_models_data:
            all_rows.extend(
                build_detailed_table(all_models_data[model_dir], model_dir, metric)
            )

        if not all_rows:
            current_row += 1
            continue

        # Column headers
        columns = list(all_rows[0].keys())
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        current_row += 1

        # Data rows
        data_start_row = current_row
        for row_data in all_rows:
            for col_idx, col_name in enumerate(columns, 1):
                val = row_data.get(col_name)
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center") if col_idx > 1 else Alignment()
                # Medium border at group boundaries
                if col_idx in (6, 8, 10):
                    cell.border = GROUP_RIGHT_BORDER
            current_row += 1

        # Bold best / underline second best (skip Model/Modality and Avg columns)
        value_columns = columns[1:]  # skip "Model / Modality"
        _apply_bold_underline(ws, all_rows, value_columns, data_start_row, col_offset=0)

        # Blank row between tables
        current_row += 2

    # Column widths
    col_widths = {"A": 35}
    for i in range(2, 12):
        col_widths[get_column_letter(i)] = 14
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = "B4"


def write_summary_sheet(ws, all_models_data: dict):
    """Write the summary table on the Summary sheet."""
    ws.cell(row=1, column=1, value="Table 4: Summary across Models and Modalities").font = TABLE_TITLE_FONT

    all_rows = []
    for model_dir in all_models_data:
        all_rows.extend(build_summary_table(all_models_data[model_dir], model_dir))

    if not all_rows:
        return

    columns = list(all_rows[0].keys())

    # Header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Data
    data_start_row = 3
    for row_idx, row_data in enumerate(all_rows):
        for col_idx, col_name in enumerate(columns, 1):
            val = row_data.get(col_name)
            cell = ws.cell(row=data_start_row + row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center") if col_idx > 1 else Alignment()

    # Bold best / underline second best for Completion column
    _apply_bold_underline(ws, all_rows, ["Completion (%)"], data_start_row, col_offset=0)

    # Column widths
    ws.column_dimensions["A"].width = 35
    for i in range(2, len(columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    ws.freeze_panes = "B3"


# --- Main ---


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    models = discover_models()

    if not models:
        print("No result CSVs found in experiment_results/")
        return

    # Load and enrich data for all models
    all_models_data = {}
    for model_dir in models:
        display = MODEL_DISPLAY.get(model_dir, model_dir)
        print(f"Loading {display} ({model_dir})...")
        try:
            df = load_result_csvs(model_dir)
            df = enrich_with_session_data(df, model_dir)
            all_models_data[model_dir] = df
        except FileNotFoundError as e:
            print(f"  Skipping: {e}")

    # Write one XLSX per model
    for model_dir in all_models_data:
        display = MODEL_DISPLAY.get(model_dir, model_dir)
        single = {model_dir: all_models_data[model_dir]}
        output_path = OUTPUT_DIR / f"{model_dir}_results.xlsx"

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Create placeholder sheets so we can access the workbook
            pd.DataFrame().to_excel(writer, sheet_name="Detailed", index=False)
            pd.DataFrame().to_excel(writer, sheet_name="Summary", index=False)
            wb = writer.book

            # Remove the auto-created empty content and write our formatted tables
            ws_detail = wb["Detailed"]
            write_detailed_sheet(ws_detail, single)

            ws_summary = wb["Summary"]
            write_summary_sheet(ws_summary, single)

        print(f"  Written: {output_path}")

    # Also write a combined XLSX with all models
    if len(all_models_data) > 1:
        output_path = OUTPUT_DIR / "all_models_results.xlsx"
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            pd.DataFrame().to_excel(writer, sheet_name="Detailed", index=False)
            pd.DataFrame().to_excel(writer, sheet_name="Summary", index=False)
            wb = writer.book

            ws_detail = wb["Detailed"]
            write_detailed_sheet(ws_detail, all_models_data)

            ws_summary = wb["Summary"]
            write_summary_sheet(ws_summary, all_models_data)

        print(f"  Written: {output_path}")

    # Print summary to console
    print("\n--- Summary ---")
    for model_dir, df in all_models_data.items():
        display = MODEL_DISPLAY.get(model_dir, model_dir)
        for method, cfg in METHOD_CONFIG.items():
            mdf = df[df["rule_adaptation_method"] == method]
            if mdf.empty:
                continue
            total = len(mdf)
            correct = mdf["all_correct"].sum()
            t = mdf["total_time_s"].dropna()
            s = mdf["ai_steps"].dropna()
            print(
                f"{display} {cfg['display']:30s}  "
                f"Completion: {correct/total*100:5.1f}%  "
                f"Time: {t.mean():6.1f} \u00b1 {t.std():5.1f}s  "
                f"Steps: {s.mean():5.1f} \u00b1 {s.std():4.1f}"
            )


if __name__ == "__main__":
    main()
